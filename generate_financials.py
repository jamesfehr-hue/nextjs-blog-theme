"""
Generate a 15-year integrated financial model with:
- Income Statement
- Balance Sheet
- Cash Flow Statement (indirect method)

All three tabs are interrelated via Excel cell references.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_SECTION_BG  = "D6E4F0"   # light blue
CLR_SECTION_FG  = "1F3864"
CLR_INPUT_BG    = "EBF5FB"   # very light blue  (hard-coded assumptions)
CLR_TOTAL_BG    = "2E86C1"   # medium blue
CLR_TOTAL_FG    = "FFFFFF"
CLR_SUBTOTAL_BG = "AED6F1"
CLR_ALT_ROW     = "F2F9FF"
CLR_BORDER      = "2E86C1"

YEARS = list(range(2024, 2039))   # 2024 – 2038 (15 years)
N     = len(YEARS)                # 15

# ── Helpers ──────────────────────────────────────────────────────────────────
def col(offset):
    """Return column letter for data columns (B = year 0, …)."""
    return get_column_letter(2 + offset)   # B, C, D, …

def thin_border(top=False, bottom=False, left=False, right=False):
    s = Side(style="thin", color=CLR_BORDER)
    n = Side(style=None)
    return Border(
        top    = s if top    else n,
        bottom = s if bottom else n,
        left   = s if left   else n,
        right  = s if right  else n,
    )

def apply_header(ws, row, label, colspan):
    """Dark header band across all columns."""
    ws.cell(row=row, column=1, value=label).font = Font(
        bold=True, color=CLR_HEADER_FG, size=11)
    ws.cell(row=row, column=1).fill = PatternFill(
        "solid", fgColor=CLR_HEADER_BG)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left",
                                                      vertical="center")
    for c in range(1, colspan + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
        cell.border = thin_border(bottom=True)

def apply_section(ws, row, label, colspan):
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = Font(bold=True, color=CLR_SECTION_FG,
                                            size=10)
    for c in range(1, colspan + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=CLR_SECTION_BG)
        cell.border = thin_border(top=True, bottom=True)

def apply_total(ws, row, label, formula_template, colspan, light=False):
    bg = CLR_SUBTOTAL_BG if light else CLR_TOTAL_BG
    fg = CLR_SECTION_FG  if light else CLR_TOTAL_FG
    ws.cell(row=row, column=1, value=label).font = Font(
        bold=True, color=fg)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=bg)
    for i in range(N):
        c  = 2 + i
        formula = formula_template.replace("{C}", get_column_letter(c)) \
                                   .replace("{R}", str(row))
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = '#,##0'
        cell.font = Font(bold=True, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin_border(top=True, bottom=True)

def write_label_row(ws, row, label, indent=0, alt=False):
    ws.cell(row=row, column=1, value=(" " * indent * 2) + label)
    ws.cell(row=row, column=1).font = Font(size=10)
    if alt:
        ws.cell(row=row, column=1).fill = PatternFill(
            "solid", fgColor=CLR_ALT_ROW)

def write_value_row(ws, row, values_or_formulas, num_fmt='#,##0', alt=False):
    bg = PatternFill("solid", fgColor=CLR_ALT_ROW) if alt else None
    for i, v in enumerate(values_or_formulas):
        cell = ws.cell(row=row, column=2 + i, value=v)
        cell.number_format = num_fmt
        cell.font = Font(size=10)
        if alt and bg:
            cell.fill = bg

def set_col_widths(ws, label_w=36, data_w=14):
    ws.column_dimensions["A"].width = label_w
    for i in range(N):
        ws.column_dimensions[get_column_letter(2 + i)].width = data_w

def write_year_headers(ws, row):
    apply_header(ws, row, "USD (thousands)", N + 1)
    ws.cell(row=row, column=1, value="USD (thousands)")
    for i, yr in enumerate(YEARS):
        c = ws.cell(row=row, column=2 + i, value=yr)
        c.font      = Font(bold=True, color=CLR_HEADER_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        c.alignment = Alignment(horizontal="center")

def freeze(ws, row=2, col_idx=2):
    ws.freeze_panes = ws.cell(row=row, column=col_idx)

# ── Base Assumptions (hard-coded into Income Statement) ──────────────────────
# Revenue grows at 8% per year; margins are stable assumptions.
REVENUE_BASE   = 50_000       # Year-1 revenue ($k)
REVENUE_GROWTH = 0.08         # 8% per year
COGS_PCT       = 0.60         # 60% of revenue
RD_PCT         = 0.08
SGA_PCT        = 0.12
DA_PCT         = 0.05         # Depreciation & Amortisation (% of revenue)
INT_RATE       = 0.05         # interest rate on debt
TAX_RATE       = 0.21
CAPEX_PCT      = 0.07         # CapEx as % of revenue
NWC_PCT        = 0.10         # Net Working Capital as % of revenue
DEBT_REPAY     = 1_000        # Annual debt repayment ($k)
DIVID_PCT      = 0.30         # Dividends as % of net income
INIT_CASH      = 5_000        # Beginning cash
INIT_DEBT      = 20_000       # Beginning LT debt
INIT_EQUITY    = 25_000       # Beginning equity (paid-in capital stays flat)
INIT_PP_E      = 15_000       # Beginning PP&E (net)

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 – INCOME STATEMENT
# ═══════════════════════════════════════════════════════════════════════════════
ws_is = wb.active
ws_is.title = "Income Statement"
ws_is.sheet_properties.tabColor = "1F3864"
set_col_widths(ws_is)
freeze(ws_is, row=3, col_idx=2)
ws_is.row_dimensions[1].height = 6

write_year_headers(ws_is, 2)

ROW = {}   # track important row numbers

# ── Revenue ──
apply_section(ws_is, 3, "REVENUE", N + 1)
ROW["IS_rev"] = 4
write_label_row(ws_is, 4, "  Net Revenue", indent=1)
rev_vals = []
for i in range(N):
    if i == 0:
        rev_vals.append(REVENUE_BASE)
    else:
        c_prev = get_column_letter(2 + i - 1)
        rev_vals.append(f"={c_prev}4*(1+{REVENUE_GROWTH})")
write_value_row(ws_is, 4, rev_vals)
# Mark input cells
ws_is.cell(4, 2).fill = PatternFill("solid", fgColor=CLR_INPUT_BG)

# ── COGS ──
apply_section(ws_is, 5, "COST OF GOODS SOLD", N + 1)
ROW["IS_cogs"] = 6
write_label_row(ws_is, 6, "  Cost of Goods Sold", indent=1, alt=True)
write_value_row(ws_is, 6,
    [f"=-{col(i)}4*{COGS_PCT}" for i in range(N)], alt=True)

apply_total(ws_is, 7, "Gross Profit",
    "={C}4+{C}6", N + 1, light=True)
ROW["IS_gp"] = 7

# ── Operating Expenses ──
apply_section(ws_is, 8, "OPERATING EXPENSES", N + 1)
ROW["IS_rd"]  = 9
ROW["IS_sga"] = 10
ROW["IS_da"]  = 11

write_label_row(ws_is, 9,  "  Research & Development",   indent=1)
write_label_row(ws_is, 10, "  Selling, General & Admin",  indent=1, alt=True)
write_label_row(ws_is, 11, "  Depreciation & Amortisation", indent=1)

write_value_row(ws_is, 9,
    [f"=-{col(i)}4*{RD_PCT}"  for i in range(N)])
write_value_row(ws_is, 10,
    [f"=-{col(i)}4*{SGA_PCT}" for i in range(N)], alt=True)
write_value_row(ws_is, 11,
    [f"=-{col(i)}4*{DA_PCT}"  for i in range(N)])

apply_total(ws_is, 12, "Total Operating Expenses",
    "=SUM({C}9:{C}11)", N + 1, light=True)
ROW["IS_totopex"] = 12

apply_total(ws_is, 13, "Operating Income (EBIT)",
    "={C}7+{C}12", N + 1)
ROW["IS_ebit"] = 13

# ── Interest & Tax ──
apply_section(ws_is, 14, "BELOW THE LINE", N + 1)
ROW["IS_int"]  = 15
ROW["IS_ebt"]  = 16
ROW["IS_tax"]  = 17
ROW["IS_ni"]   = 18

write_label_row(ws_is, 15, "  Interest Expense", indent=1)
# Interest references BS debt (handled after BS is built – use placeholder formula)
# For IS we just reference "Balance Sheet"!debt row (built below)
# We'll use a forward reference: BS long-term debt is at row 9 (to be confirmed)
write_value_row(ws_is, 15,
    [f"=-'Balance Sheet'!{col(i)}9*{INT_RATE}" for i in range(N)])

apply_total(ws_is, 16, "Earnings Before Tax (EBT)",
    "={C}13+{C}15", N + 1, light=True)
write_value_row(ws_is, 17,
    [f"=-MAX({col(i)}16*{TAX_RATE},0)" for i in range(N)])
write_label_row(ws_is, 17, "  Income Tax Expense", indent=1)

apply_total(ws_is, 18, "Net Income",
    "={C}16+{C}17", N + 1)

# ── EPS / Margins (informational) ──
apply_section(ws_is, 19, "MARGINS", N + 1)
write_label_row(ws_is, 20, "  Gross Margin %",   indent=1)
write_label_row(ws_is, 21, "  EBIT Margin %",    indent=1, alt=True)
write_label_row(ws_is, 22, "  Net Profit Margin %", indent=1)
for i in range(N):
    c = col(i)
    ws_is.cell(20, 2+i, value=f"={c}7/{c}4").number_format  = "0.0%"
    ws_is.cell(21, 2+i, value=f"={c}13/{c}4").number_format = "0.0%"
    ws_is.cell(22, 2+i, value=f"={c}18/{c}4").number_format = "0.0%"
    ws_is.cell(21, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 – BALANCE SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws_bs = wb.create_sheet("Balance Sheet")
ws_bs.sheet_properties.tabColor = "117A65"
set_col_widths(ws_bs)
freeze(ws_bs, row=3, col_idx=2)
ws_bs.row_dimensions[1].height = 6

write_year_headers(ws_bs, 2)

# ── ASSETS ──
apply_section(ws_bs, 3, "CURRENT ASSETS", N + 1)
ROW["BS_cash"]  = 4
ROW["BS_ar"]    = 5
ROW["BS_inv"]   = 6
ROW["BS_ca"]    = 7

write_label_row(ws_bs, 4, "  Cash & Equivalents",   indent=1)
write_label_row(ws_bs, 5, "  Accounts Receivable",  indent=1, alt=True)
write_label_row(ws_bs, 6, "  Inventory",             indent=1)

# Cash pulled from Cash Flow Statement ending balance
for i in range(N):
    c = col(i)
    ws_bs.cell(4, 2+i, value=f"='Cash Flow'!{c}22").number_format = '#,##0'

# AR & Inventory driven by NWC % of revenue
for i in range(N):
    c = col(i)
    ws_bs.cell(5, 2+i, value=f"='Income Statement'!{c}4*0.08") \
         .number_format = '#,##0'
    ws_bs.cell(5, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)
    ws_bs.cell(6, 2+i, value=f"='Income Statement'!{c}4*0.06") \
         .number_format = '#,##0'

apply_total(ws_bs, 7, "Total Current Assets",
    "=SUM({C}4:{C}6)", N + 1, light=True)

# Non-Current Assets
apply_section(ws_bs, 8, "NON-CURRENT ASSETS", N + 1)
ROW["BS_ppe"] = 9

write_label_row(ws_bs, 9, "  Property, Plant & Equipment (net)", indent=1)
for i in range(N):
    c     = col(i)
    c_prev = get_column_letter(2 + i - 1)
    if i == 0:
        ws_bs.cell(9, 2+i,
            value=f"={INIT_PP_E}+'Cash Flow'!{c}18-'Income Statement'!{c}11"
        ).number_format = '#,##0'
        ws_bs.cell(9, 2).fill = PatternFill("solid", fgColor=CLR_INPUT_BG)
    else:
        ws_bs.cell(9, 2+i,
            value=f"={c_prev}9+'Cash Flow'!{c}18-'Income Statement'!{c}11"
        ).number_format = '#,##0'

write_label_row(ws_bs, 10, "  Intangibles & Goodwill", indent=1, alt=True)
for i in range(N):
    ws_bs.cell(10, 2+i, value=2000).number_format = '#,##0'
    ws_bs.cell(10, 2+i).fill = PatternFill("solid", fgColor=CLR_INPUT_BG)

apply_total(ws_bs, 11, "Total Non-Current Assets",
    "=SUM({C}9:{C}10)", N + 1, light=True)

apply_total(ws_bs, 12, "TOTAL ASSETS",
    "={C}7+{C}11", N + 1)
ROW["BS_ta"] = 12

# ── LIABILITIES ──
apply_section(ws_bs, 14, "CURRENT LIABILITIES", N + 1)
ROW["BS_ap"]  = 15
ROW["BS_accr"]= 16
ROW["BS_cl"]  = 17

write_label_row(ws_bs, 15, "  Accounts Payable",  indent=1)
write_label_row(ws_bs, 16, "  Accrued Liabilities", indent=1, alt=True)
for i in range(N):
    c = col(i)
    ws_bs.cell(15, 2+i,
        value=f"='Income Statement'!{c}6*(-0.25)").number_format = '#,##0'
    ws_bs.cell(16, 2+i,
        value=f"='Income Statement'!{c}4*0.03").number_format = '#,##0'
    ws_bs.cell(16, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

apply_total(ws_bs, 17, "Total Current Liabilities",
    "=SUM({C}15:{C}16)", N + 1, light=True)

apply_section(ws_bs, 18, "NON-CURRENT LIABILITIES", N + 1)
ROW["BS_debt"] = 19
write_label_row(ws_bs, 19, "  Long-Term Debt", indent=1)
for i in range(N):
    c_prev = get_column_letter(2 + i - 1)
    if i == 0:
        ws_bs.cell(19, 2+i, value=INIT_DEBT - DEBT_REPAY).number_format='#,##0'
        ws_bs.cell(19, 2).fill = PatternFill("solid", fgColor=CLR_INPUT_BG)
    else:
        ws_bs.cell(19, 2+i,
            value=f"=MAX({c_prev}19-{DEBT_REPAY},0)").number_format = '#,##0'

apply_total(ws_bs, 20, "Total Non-Current Liabilities",
    "={C}19", N + 1, light=True)
apply_total(ws_bs, 21, "TOTAL LIABILITIES",
    "={C}17+{C}20", N + 1)
ROW["BS_tl"] = 21

# ── EQUITY ──
apply_section(ws_bs, 23, "SHAREHOLDERS' EQUITY", N + 1)
ROW["BS_paidin"]  = 24
ROW["BS_re"]      = 25
ROW["BS_te"]      = 26

write_label_row(ws_bs, 24, "  Paid-In Capital",    indent=1)
write_label_row(ws_bs, 25, "  Retained Earnings",  indent=1, alt=True)

for i in range(N):
    ws_bs.cell(24, 2+i, value=INIT_EQUITY).number_format = '#,##0'
    ws_bs.cell(24, 2+i).fill = PatternFill("solid", fgColor=CLR_INPUT_BG)

for i in range(N):
    c      = col(i)
    c_prev = get_column_letter(2 + i - 1)
    if i == 0:
        ws_bs.cell(25, 2+i,
            value=(f"='Income Statement'!{c}18"
                   f"*(1-{DIVID_PCT})")
        ).number_format = '#,##0'
    else:
        ws_bs.cell(25, 2+i,
            value=(f"={c_prev}25+'Income Statement'!{c}18"
                   f"*(1-{DIVID_PCT})")
        ).number_format = '#,##0'
    ws_bs.cell(25, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

apply_total(ws_bs, 26, "TOTAL EQUITY",
    "={C}24+{C}25", N + 1)

apply_total(ws_bs, 28, "TOTAL LIABILITIES & EQUITY",
    "={C}21+{C}26", N + 1)
ROW["BS_tle"] = 28

# Balance Check row
apply_section(ws_bs, 29, "BALANCE CHECK (Assets = Liabilities + Equity)", N + 1)
for i in range(N):
    c    = col(i)
    cell = ws_bs.cell(30, 2+i,
        value=f"=IF(ABS({c}12-{c}28)<1,\"OK\",\"MISMATCH\")")
    cell.font      = Font(bold=True, color="117A65")
    cell.alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 – CASH FLOW STATEMENT (indirect method)
# ═══════════════════════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet("Cash Flow")
ws_cf.sheet_properties.tabColor = "7B241C"
set_col_widths(ws_cf)
freeze(ws_cf, row=3, col_idx=2)
ws_cf.row_dimensions[1].height = 6

write_year_headers(ws_cf, 2)

# ── Operating Activities ──
apply_section(ws_cf, 3, "OPERATING ACTIVITIES", N + 1)

write_label_row(ws_cf, 4,  "  Net Income",             indent=1)
write_label_row(ws_cf, 5,  "  Add: D&A",               indent=1, alt=True)
write_label_row(ws_cf, 6,  "  Changes in Working Capital", indent=1)
write_label_row(ws_cf, 7,  "    Δ Accounts Receivable",  indent=2, alt=True)
write_label_row(ws_cf, 8,  "    Δ Inventory",            indent=2)
write_label_row(ws_cf, 9,  "    Δ Accounts Payable",     indent=2, alt=True)
write_label_row(ws_cf, 10, "    Δ Accrued Liabilities",  indent=2)

for i in range(N):
    c      = col(i)
    c_prev = get_column_letter(2 + i - 1)

    ws_cf.cell(4, 2+i,
        value=f"='Income Statement'!{c}18").number_format = '#,##0'

    ws_cf.cell(5, 2+i,
        value=f"=-'Income Statement'!{c}11").number_format = '#,##0'
    ws_cf.cell(5, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

    if i == 0:
        ws_cf.cell(7,  2+i, value=0).number_format = '#,##0'
        ws_cf.cell(8,  2+i, value=0).number_format = '#,##0'
        ws_cf.cell(9,  2+i, value=0).number_format = '#,##0'
        ws_cf.cell(10, 2+i, value=0).number_format = '#,##0'
    else:
        ws_cf.cell(7, 2+i,
            value=f"=-('Balance Sheet'!{c}5-'Balance Sheet'!{c_prev}5)"
        ).number_format = '#,##0'
        ws_cf.cell(8, 2+i,
            value=f"=-('Balance Sheet'!{c}6-'Balance Sheet'!{c_prev}6)"
        ).number_format = '#,##0'
        ws_cf.cell(9, 2+i,
            value=f"='Balance Sheet'!{c}15-'Balance Sheet'!{c_prev}15"
        ).number_format = '#,##0'
        ws_cf.cell(10, 2+i,
            value=f"='Balance Sheet'!{c}16-'Balance Sheet'!{c_prev}16"
        ).number_format = '#,##0'

    ws_cf.cell(7,  2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)
    ws_cf.cell(9,  2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

apply_total(ws_cf, 11, "Total Changes in Working Capital",
    "=SUM({C}7:{C}10)", N + 1, light=True)
apply_total(ws_cf, 12, "Cash from Operations",
    "=SUM({C}4:{C}5)+{C}11", N + 1)
ROW["CF_ops"] = 12

# ── Investing Activities ──
apply_section(ws_cf, 14, "INVESTING ACTIVITIES", N + 1)
ROW["CF_capex"] = 15
write_label_row(ws_cf, 15, "  Capital Expenditures (CapEx)", indent=1)
for i in range(N):
    c = col(i)
    ws_cf.cell(15, 2+i,
        value=f"=-'Income Statement'!{c}4*{CAPEX_PCT}"
    ).number_format = '#,##0'

apply_total(ws_cf, 16, "Cash from Investing",
    "={C}15", N + 1)
ROW["CF_inv"] = 16

# ── Financing Activities ──
apply_section(ws_cf, 18, "FINANCING ACTIVITIES", N + 1)
ROW["CF_debt_pay"] = 19
ROW["CF_div"]      = 20
write_label_row(ws_cf, 19, "  Debt Repayment",       indent=1)
write_label_row(ws_cf, 20, "  Dividends Paid",        indent=1, alt=True)

for i in range(N):
    c = col(i)
    ws_cf.cell(19, 2+i, value=-DEBT_REPAY).number_format = '#,##0'
    ws_cf.cell(20, 2+i,
        value=f"=-'Income Statement'!{c}18*{DIVID_PCT}"
    ).number_format = '#,##0'
    ws_cf.cell(20, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

apply_total(ws_cf, 21, "Cash from Financing",
    "=SUM({C}19:{C}20)", N + 1)
ROW["CF_fin"] = 21

# ── Net Change in Cash / Ending Balance ──
apply_section(ws_cf, 23, "NET CASH POSITION", N + 1)
ROW["CF_beg_cash"] = 24
ROW["CF_net_chg"]  = 25
ROW["CF_end_cash"] = 26   # → BS cash row 4

write_label_row(ws_cf, 24, "  Beginning Cash Balance", indent=1)
write_label_row(ws_cf, 25, "  Net Change in Cash",     indent=1, alt=True)

for i in range(N):
    c      = col(i)
    c_prev = get_column_letter(2 + i - 1)
    if i == 0:
        ws_cf.cell(24, 2+i, value=INIT_CASH).number_format = '#,##0'
        ws_cf.cell(24, 2).fill = PatternFill("solid", fgColor=CLR_INPUT_BG)
    else:
        ws_cf.cell(24, 2+i,
            value=f"={c_prev}22").number_format = '#,##0'

    ws_cf.cell(25, 2+i,
        value=f"={c}12+{c}16+{c}21").number_format = '#,##0'
    ws_cf.cell(25, 2+i).fill = PatternFill("solid", fgColor=CLR_ALT_ROW)

apply_total(ws_cf, 22, "Ending Cash Balance",
    "={C}24+{C}25", N + 1)
# re-label row 22 (already written by apply_total via row number)

# ── Debt Reference row (so BS interest ref works) ──
# BS row 9 = debt – this is already built above at ws_bs row 19
# IS interest row 15 references 'Balance Sheet'!col(i)9 — BUT that's PP&E!
# Fix: IS interest should reference BS row 19 (LT Debt), not row 9.
# Patch the interest formulas in IS:
for i in range(N):
    c = col(i)
    c_prev = get_column_letter(2 + i - 1) if i > 0 else col(0)
    # Use beginning-of-period debt for interest calc
    if i == 0:
        ws_is.cell(15, 2+i,
            value=f"=-{INIT_DEBT}*{INT_RATE}").number_format = '#,##0'
    else:
        ws_is.cell(15, 2+i,
            value=f"=-'Balance Sheet'!{c_prev}19*{INT_RATE}"
        ).number_format = '#,##0'

# ── Summary tab ──────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Summary", 0)
wb.active = ws_sum
ws_sum.sheet_properties.tabColor = "6C3483"
set_col_widths(ws_sum, label_w=40)
ws_sum.row_dimensions[1].height = 20
ws_sum.row_dimensions[2].height = 30

# Title
ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N + 1)
tc = ws_sum.cell(1, 1, value="15-YEAR INTEGRATED FINANCIAL MODEL")
tc.font      = Font(bold=True, size=16, color=CLR_HEADER_FG)
tc.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
tc.alignment = Alignment(horizontal="center", vertical="center")

write_year_headers(ws_sum, 2)

metrics = [
    ("Revenue",             "='Income Statement'!{C}4"),
    ("Gross Profit",        "='Income Statement'!{C}7"),
    ("EBIT",                "='Income Statement'!{C}13"),
    ("Net Income",          "='Income Statement'!{C}18"),
    ("Cash from Operations","='Cash Flow'!{C}12"),
    ("CapEx",               "='Cash Flow'!{C}15"),
    ("Free Cash Flow",      "='Cash Flow'!{C}12+'Cash Flow'!{C}15"),
    ("Ending Cash",         "='Cash Flow'!{C}22"),
    ("Total Assets",        "='Balance Sheet'!{C}12"),
    ("Total Debt",          "='Balance Sheet'!{C}19"),
    ("Total Equity",        "='Balance Sheet'!{C}26"),
    ("Gross Margin %",      "='Income Statement'!{C}20"),
    ("Net Margin %",        "='Income Statement'!{C}22"),
    ("Debt / Equity",       "='Balance Sheet'!{C}19/'Balance Sheet'!{C}26"),
]

for r_off, (label, formula) in enumerate(metrics):
    row = 3 + r_off
    alt = r_off % 2 == 1
    write_label_row(ws_sum, row, label, alt=alt)
    for i in range(N):
        c    = col(i)
        f    = formula.replace("{C}", c)
        cell = ws_sum.cell(row, 2 + i, value=f)
        cell.number_format = "0.0%" if "%" in label or "Ratio" in label or "/" in label else '#,##0'
        if alt:
            cell.fill = PatternFill("solid", fgColor=CLR_ALT_ROW)
        cell.font = Font(size=10)

freeze(ws_sum, row=3, col_idx=2)
set_col_widths(ws_sum)

# ── Save ─────────────────────────────────────────────────────────────────────
out_path = "/home/user/nextjs-blog-theme/financial_model_15yr.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
